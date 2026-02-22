from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ..compare.workflow import load_compare_run_data
from ..storage.sqlite import SQLiteStore


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _write_csv(path: Path, headers: list[str], rows: list[dict[str, object]]) -> Path:
    _ensure_parent(path)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in headers})
    return path


def _write_xlsx(
    path: Path,
    *,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, object]],
    metadata: dict[str, object],
) -> Path:
    _ensure_parent(path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    meta_ws = wb.create_sheet("metadata")
    meta_ws.append(["key", "value"])
    meta_ws["A1"].font = Font(bold=True)
    meta_ws["B1"].font = Font(bold=True)
    for key, value in metadata.items():
        if isinstance(value, dict | list):
            value_text = json.dumps(value, ensure_ascii=False, sort_keys=True)
        else:
            value_text = str(value)
        meta_ws.append([key, value_text])

    wb.save(path)
    return path


def export_compare_coverage(store: SQLiteStore, run_id: str, out_path: Path) -> Path:
    payload = load_compare_run_data(store.db_path, run_id)
    coverage = payload["coverage"]
    corpus_ids = [str(corpus_id) for corpus_id in coverage.get("corpus_ids", [])]

    headers = ["headword_key", *[f"in_{corpus_id}" for corpus_id in corpus_ids]]
    rows: list[dict[str, object]] = []
    for row in coverage.get("rows", []):
        row_dict = {"headword_key": row.get("headword_key", "")}
        presence = row.get("presence", {})
        if not isinstance(presence, dict):
            presence = {}
        for corpus_id in corpus_ids:
            row_dict[f"in_{corpus_id}"] = "1" if bool(presence.get(corpus_id, False)) else "0"
        rows.append(row_dict)

    metadata = {
        "run_id": run_id,
        "corpus_ids": corpus_ids,
        "counts": coverage.get("counts", {}),
    }

    if out_path.suffix.lower() == ".xlsx":
        return _write_xlsx(
            out_path,
            sheet_name="coverage",
            headers=headers,
            rows=rows,
            metadata=metadata,
        )
    return _write_csv(out_path, headers=headers, rows=rows)


def export_compare_alignment(store: SQLiteStore, run_id: str, out_path: Path) -> Path:
    payload = load_compare_run_data(store.db_path, run_id)
    alignment = payload["alignment"]

    headers = [
        "headword_key",
        "headword_a",
        "headword_b",
        "headword_norm_a",
        "headword_norm_b",
        "entry_id_a",
        "entry_id_b",
        "status_a",
        "status_b",
        "score",
        "method",
        "reason",
    ]
    rows = [
        {
            "headword_key": row.get("headword_key", ""),
            "headword_a": row.get("headword_a", ""),
            "headword_b": row.get("headword_b", ""),
            "headword_norm_a": row.get("headword_norm_a", ""),
            "headword_norm_b": row.get("headword_norm_b", ""),
            "entry_id_a": row.get("entry_id_a", ""),
            "entry_id_b": row.get("entry_id_b", ""),
            "status_a": row.get("status_a", ""),
            "status_b": row.get("status_b", ""),
            "score": row.get("score", 0),
            "method": row.get("method", ""),
            "reason": row.get("reason", ""),
        }
        for row in alignment.get("rows", [])
    ]

    metadata = {
        "run_id": run_id,
        "counts": alignment.get("counts", {}),
    }

    if out_path.suffix.lower() == ".xlsx":
        return _write_xlsx(
            out_path,
            sheet_name="alignment",
            headers=headers,
            rows=rows,
            metadata=metadata,
        )
    return _write_csv(out_path, headers=headers, rows=rows)


def export_compare_diff(store: SQLiteStore, run_id: str, out_path: Path) -> Path:
    payload = load_compare_run_data(store.db_path, run_id)
    diff = payload["diff"]

    headers = [
        "headword_key",
        "entry_id_a",
        "entry_id_b",
        "pron_render_a",
        "pron_render_b",
        "pron_norm_a",
        "pron_norm_b",
        "features_a_json",
        "features_b_json",
        "delta_json",
    ]
    rows = [
        {
            "headword_key": row.get("headword_key", ""),
            "entry_id_a": row.get("entry_id_a", ""),
            "entry_id_b": row.get("entry_id_b", ""),
            "pron_render_a": row.get("pron_render_a", ""),
            "pron_render_b": row.get("pron_render_b", ""),
            "pron_norm_a": row.get("pron_norm_a", ""),
            "pron_norm_b": row.get("pron_norm_b", ""),
            "features_a_json": json.dumps(
                row.get("features_a", {}), ensure_ascii=False, sort_keys=True
            ),
            "features_b_json": json.dumps(
                row.get("features_b", {}), ensure_ascii=False, sort_keys=True
            ),
            "delta_json": json.dumps(row.get("delta", {}), ensure_ascii=False, sort_keys=True),
        }
        for row in diff.get("rows", [])
    ]

    metadata = {
        "run_id": run_id,
        "counts": diff.get("counts", {}),
    }

    if out_path.suffix.lower() == ".xlsx":
        return _write_xlsx(
            out_path,
            sheet_name="diff",
            headers=headers,
            rows=rows,
            metadata=metadata,
        )
    return _write_csv(out_path, headers=headers, rows=rows)

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ..storage.sqlite import SQLiteStore


def export_comparison_xlsx(store: SQLiteStore, dict_ids: list[str], out_path: Path) -> Path:
    rows = store.comparison_rows(dict_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "matrix"

    headers = ["lemma_group_id", "lemma_label", *dict_ids]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        values = row.get("values", {})
        ws.append(
            [
                row["lemma_group_id"],
                row["lemma_label"],
                *[values.get(d, "ABSENT") for d in dict_ids],
            ]
        )

    meta = wb.create_sheet(title="metadata_flags")
    meta.append(["metric", "value"])
    meta["A1"].font = Font(bold=True)
    meta["B1"].font = Font(bold=True)
    meta.append(["dictionaries", ", ".join(dict_ids)])
    meta.append(["lemma_groups", len(rows)])
    for top in store.top_issues(limit=30):
        meta.append([f"issue:{top['code']}:{top['kind']}", int(top["n"])])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
